import csv
from pathlib import Path
import random
import time
from get_champions import get_champions
from get_champions_counters import get_champion_counters
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

'''
GET LIST OF CHAMPIONS WITH RATES AND WRITE IN CSV
'''
def write_csv_champions(lane, tier = "emerald_plus", patch = None):
    champions = get_champions(lane, tier=tier, patch=patch)
    rows = list()
    for champ in champions:
        name = list()
        name.append(champ[0])
        name.append(champ[1])
        name.append(champ[2])
        name.append(champ[3])
        rows.append(name)
    with open("champions_{}_{}.csv".format(lane, tier), "w", newline='') as f: 
        write = csv.writer(f) 
        write.writerows(rows) 
    

'''
READ LIST OF CHAMPIONS WITH RATES FROM CSV
'''
def read_csv_champions(lane, tier = "emerald_plus"):
    file = open("champions_{}_{}.csv".format(lane, tier), "r")
    champions = list(csv.reader(file, delimiter=","))
    file.close()
    result = list()
    for champ in champions:
        result.append((champ[0], champ[1], champ[2], champ[3]))
    return result



'''
READ A CHAMPION MATCHUP / COUNTERS AND WRITE IN CSV
'''
def write_csv_champion_counters(champion, lane, tier = "emerald_plus", patch = None):
    counters = get_champion_counters(champion, lane, tier=tier, patch=patch)
    rows = list()
    file_path = "matchups/{}/{}/counters_{}.csv".format(lane, tier, champion)
    folder_path = "matchups/{}/{}/".format(lane, tier)
    for champ in counters:
        row = list()        
        row.append(champ)
        row.append(counters[champ][0])
        row.append(counters[champ][1])
        rows.append(row)
    Path(folder_path).mkdir(parents=True, exist_ok=True)
    with open(file_path, "w", newline='') as f: 
        write = csv.writer(f) 
        write.writerows(rows) 


'''
READ A CHAMPION MATCHUP / COUNTERS FROM CSV
'''
def read_csv_champion_counters(champion, lane, tier = "emerald_plus"):
    file = open("matchups/{}/{}/counters_{}.csv".format(lane, tier, champion), "r")
    counters = list(csv.reader(file, delimiter=","))
    file.close()
    result = dict()
    for champ in counters:
        result[champ[0]] = (champ[1], champ[2])
    return result


'''
READ CHAMPION RATES + MATCHUPS AND WRITE A CSV WITH GLOBAL MATCHUP INFO
'''
def write_csv_all_lane_matchups(lane, tier = "emerald_plus"):
    champions = read_csv_champions(lane, tier=tier)
    champions_formatted = [x[0].replace(".", "").replace(" ", "").replace("'", "").lower() for x in champions]
    global_counters = dict()
    for x in champions_formatted:
        counters = read_csv_champion_counters(x, lane, tier=tier)
        counters_formatted = dict()
        for y in counters:
            counters_formatted[y.replace(".", "").replace(" ", "").replace("'", "").lower()] = (float(counters[y][0].replace("%", "")), int(counters[y][1].replace(",", "")))
        global_counters[x] = counters_formatted    

    champions_formatted.sort()
    rows = list()
    for champ in champions_formatted:
        row = list()
        row.append(champ)
        counters = global_counters[champ]
        for opponent in champions_formatted:
            stat = counters.get(opponent)
            row.append((stat[0] if stat is not None else None))
        rows.append(row)

    headers = ["SELF"]
    headers = headers + champions_formatted
    with open("matchups_{}_{}_all.csv".format(lane, tier), "w", newline='') as f: 
        write = csv.writer(f) 
        write.writerow(headers)
        write.writerows(rows) 


'''
REFRESH ALL DATA : GET CHAMPIONS / COUNTERS / GLOBAL MATCHUPS AND REWRITE ALL CSVs
'''
def refresh_data(lane, tier = "emerald_plus", patch = None):
    write_csv_champions(lane, tier=tier, patch=patch)
    champions = read_csv_champions(lane, tier=tier)
    champions_formatted = [x[0].replace(".", "").replace(" ", "").replace("'", "").lower() for x in champions]
    for x in champions_formatted:
        write_csv_champion_counters(x, lane, tier=tier, patch=patch)
        time.sleep(2)
    write_csv_all_lane_matchups(lane, tier=tier)

def read_csv_all_champions_data(lane, tier = "emerald_plus"):
    result = dict()
    champions = read_csv_champions(lane, tier=tier)
    champions_formatted = [(x[0].replace(".", "").replace(" ", "").replace("'", "").lower(), float(x[1].replace("%", "")), float(x[2].replace("%", "")), float(x[3].replace("%", ""))) for x in champions]
    for x in champions_formatted:
        result[x[0]] = dict()
        result[x[0]]["winrate"] = x[1]
        result[x[0]]["pickrate"] = x[2]
        result[x[0]]["banrate"] = x[3]
        result[x[0]]["counters"] = read_csv_champion_counters(x[0], lane)
    return result


def generate_1000_matchups_score(lane = "top"):
    champions = read_csv_all_champions_data(lane)
    champions_name = [x for x in champions]
    matchs = [random.choice(champions_name) for x in range(1000)]
    print(matchs)

    result = dict()
    for x in champions_name:
        result[x] = 0
        counters = champions[x]["counters"]
        counters_formatter = { x.replace(".", "").replace(" ", "").replace("'", "").lower() : float(counters[x][0].replace("%", "")) for x in counters }
        for y in matchs:
            matchup = counters_formatter.get(y)
            if matchup is not None:
                result[x] += (matchup - 50)

    result_list = [ (x, int(result[x])) for x in result ]
    result_list.sort(key=lambda x: x[1], reverse=True)

    rows = list()
    for x in result_list:
        row = [ x[0], x[1] ]
        rows.append(row)

    with open("1000_matchs_champions_score_top.csv", "w", newline='') as f: 
        write = csv.writer(f) 
        write.writerows(rows) 



def read_csv_all_matchups(lane, tier = "emerald_plus"):
    file = open("matchups_{}_{}_all.csv".format(lane, tier), "r")
    champions = list(csv.reader(file, delimiter=","))
    file.close()
    return champions


def write_my_matchups(champions, lane, tier):
    champions = read_csv_champions(lane=lane, tier=tier)
    champions_formatter = [ x[0].replace(".", "").replace(" ", "").replace("'", "").lower() for x in champions ]
    champions_formatter.sort()
    rows = list()

    for x in champions_formatter:
        counters = read_csv_champion_counters(champion=x, lane=lane, tier=tier)
        counters_formatted = { k.replace(".", "").replace(" ", "").replace("'", "").lower() : v for k, v in counters.items() }
        my_counters = [ (k, v[0], v[1]) for k,v in counters_formatted.items() if k in my_champions ]
        my_counters_sorted = sorted(my_counters, key=lambda d: d[1])
        row = [ x ]
        for y in my_counters_sorted:
            row.extend( [ y[0], y[1].replace("%", "")] )
        rows.append(row)

    with open("my_matchups_{}_{}.csv".format(lane, tier), "w", newline='') as f: 
        write = csv.writer(f) 
        write.writerows(rows) 


def read_my_matchups(lane, tier = "emerald_plus"):
    file = open("my_matchups_{}_{}.csv".format(lane, tier), "r")
    matchups = list(csv.reader(file, delimiter=","))
    file.close()
    return matchups


def refresh_ranks(lane, ranks, patch = None, champions = None):
    for rank in ranks:
        refresh_data(lane=lane, tier=rank, patch=patch)
        write_csv_all_lane_matchups(LANE, RANK)
        write_my_matchups(champions=champions, lane=LANE, tier=RANK)
        print(f"done : lane={lane} rank={rank}, patch={patch}")

def generate_my_full_matchups_files(lane, ranks):
    book = Workbook()
    sheet = book.active
    sheet.title = "My Full Matchup"
    for rank in ranks:
        matchups = read_my_matchups(lane=lane, tier=rank)
        rank_sheet = book.create_sheet(rank)
        for matchup in matchups:
            rank_sheet.append(matchup)
    book.save(filename = "myfullmatchup.xlsx")

def colorize_my_matchups(filename, champ_colors):
    wb = load_workbook(filename)
    for sheet in wb:
        print(sheet.title)
        for row in sheet.iter_rows(min_row=1, max_col=2, max_row=sheet.max_row):
            pick_cell = row[1]
            if pick_cell.value is not None:
                champ = [ x for x in champ_colors if x[0] == pick_cell.value ]
                pick_cell.fill = champ[0][1]
    wb.save(filename)          

LANE = "top"
RANK = "iron"
PATCH = None
my_champions = [ "drmundo", "nasus", "tryndamere", "kayle", "garen", "sett", "mordekaiser", "darius", "yorick" ]

champ_colors = [ 
    ("drmundo", PatternFill(start_color="ABEBC6", end_color="ABEBC6", fill_type = "solid")),
    ("nasus", PatternFill(start_color="a569bd", end_color="a569bd", fill_type = "solid")),
    ("tryndamere", PatternFill(start_color="fadbd8", end_color="fadbd8", fill_type = "solid")),
    ("kayle", PatternFill(start_color="fcf3cf", end_color="fcf3cf", fill_type = "solid")),
    ("garen", PatternFill(start_color="f9e79f", end_color="f9e79f", fill_type = "solid")),
    ("sett", PatternFill(start_color="f5cba7", end_color="f5cba7", fill_type = "solid")),
    ("mordekaiser", PatternFill(start_color="d5dbdb", end_color="d5dbdb", fill_type = "solid")),
    ("darius", PatternFill(start_color="f5b7b1", end_color="f5b7b1", fill_type = "solid")),
    ("yorick", PatternFill(start_color="aed6f1", end_color="aed6f1", fill_type = "solid"))
]

ranks = [ "iron", "bronze", "silver", "gold", "emerald_plus", "diamond_plus", "master_plus" ] 

MATCHUP_EXCEL = "myfullmatchup.xlsx"

# refresh_ranks(lane=LANE, ranks=ranks, patch=PATCH, champions=my_champions)
# for rank in ranks:
#     write_my_matchups(champions=my_champions, lane=LANE, tier=rank)   
# generate_my_full_matchups_files(lane=LANE, ranks=ranks)

#refresh_data(lane=LANE, tier=RANK, patch=PATCH)
#write_my_matchups(champions=my_champions, lane=LANE, tier=RANK)

          
    

